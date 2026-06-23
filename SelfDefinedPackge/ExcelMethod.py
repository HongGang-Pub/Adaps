import os
import re

# import openpyxl
import json
from tkinter import messagebox


def save_excel(fname: str, sheet_name: str, data_list: list, fd_path: str, note: str = None) -> str:
    """
    根据用户要求保存文件

    Args:
        fname(str): filename
        sheet_name(str): sheet name(如果指定的文件中存在同名的sheet_name，会删除后重新写入)
        data_list(list): 存储的信息内容
        fd_path(str): 文件存储路径，若指定文件夹不存在，则创建。若不指定，则默认为当前文件夹
        note(str): 便签，不为None时，打印日志 & 文件路径；否则不打印

    Returns:
        str: 返回保存的文件路径
    """

    file = "{}\\{}.xlsx".format(fd_path, fname)
    if not os.path.exists(fd_path):
        # 目录不存在，进行创建操作
        os.makedirs(fd_path)
    state = True
    if os.path.exists(file):
        while state:
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(file)
                break
            except BaseException as msg:
                state = messagebox.askretrycancel('弹窗', '文件无法读取，请关闭文件后再试?\n{}'.format(msg))
        # os.remove(fd_path)
        if not state:
            return
    else:
        import openpyxl
        workbook = openpyxl.Workbook()
        # sheet = workbook.active
        # sheet.title = 'SpadArray_{}'.format(coeff)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)

    sheet = workbook.create_sheet(sheet_name, 0)
    for row in data_list:
        sheet.append(row)
    # alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
    # cell = sheet["A1"]
    # cell.alignment = alignment
    # sheet["A1"].font = f1
    # sheet["A1"].fill = fill1

    while state:
        try:
            workbook.save(file)
            break
        except BaseException as msg:
            state = messagebox.askretrycancel('弹窗', '文件无法读取，请关闭文件后再试?\n{}'.format(msg))
    # os.remove(fd_path)
    if not state:
        return

    if note:
        print("{}，文件路径为:  {}".format(note, file))
    return file

    # data_statistics.append("{:>3}\t{:>3}".format(r, c))


def get_bit_range(s):
    nums = re.findall(r'\d+', s)
    if len(nums) == 2:
        msb, lsb = map(int, nums)
        return msb, lsb, msb - lsb + 1  # 返回 高位, 低位, 宽度
    elif len(nums) == 1:
        bit = int(nums[0])
        return bit, bit, 1  # 返回 位点, 位点, 宽度 1
    return None


class ExcelRead(object):
    def __init__(self, file):
        self.f = file
        self.wb = None
        self.DEBUG = False
        self.get_workbook()

    def get_workbook(self):
        if self.f.split('.')[-1] == 'xlsx':
            import openpyxl
            self.wb = openpyxl.load_workbook(self.f)
        else:
            raise Exception("File format error, not excel file!")

    def get_sheets_names(self):
        if self.wb:
            return self.wb.sheetnames
        else:
            raise Exception("Workbook is None!")

    def get_register_data(self, sheet_name=None):
        """
        这个方法仅支持特定格式的 Excel 读取 register 信息
        Args:
            sheet_name:

        Returns:

        """
        register = {}
        reg_field = {}
        reg_name = None

        sheet = self.wb.worksheets[0] if sheet_name is None else self.wb[sheet_name]
        sheet_value = list(sheet.iter_rows(values_only=True))
        reg_struct = sheet_value.pop(0)  # reg_struct = [address, reg_name, bits, field, type, default_value]

        for row_value in sheet_value:
            for index in range(len(reg_struct)):
                reg_field[reg_struct[index]] = row_value[index]

            reg_name = reg_name if reg_field["reg_name"] is None else reg_field["reg_name"]
            if reg_name not in register:
                register[reg_name] = {
                    'address': reg_field["address"],
                    'reg_defalut_value': 0,
                    'field': {}
                }

            field_bit_info = get_bit_range(reg_field["bits"])
            register[reg_name]["field"][reg_field["field"]] = {
                'bits': field_bit_info,
                'type': reg_field["type"],
                'default_value': reg_field["default_value"],
                'field_width': 0,
                'field_low_bit': 0
            }

        if self.DEBUG:
            pwd = os.getcwd()
            file_name = f"{pwd}/{sheet_name}.json"
            with open(file_name, "w", encoding='utf-8') as write:
                json.dump(register, write, indent=4, ensure_ascii=False, separators=(",", ": "))
        return register


if __name__ == "__main__":
    er = ExcelRead(file=r"../REG_MODEL/reg.xlsx")
    er.DEBUG = True
    er.get_register_data()
    s = get_bit_range("[31:16]")
    print(s)
