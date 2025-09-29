#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
"""
=================================================================================================
@FileName    : Swan01ROISramOperation.py
@Author      : honggang_li
@Email       : honggang.li@adaps-ph.com

@Function    :

@Modify Time        @Author        @Version    @Description
----------------    -----------    --------    -------------
2025-07-11 14:24    honggang_li    v1.0        ROI Sram 相关操作方法

=================================================================================================
"""
import math

import openpyxl as op
from SelfDefinedPackge import PubMethod


def roi_config_dict_ini():
    roi_config_dict = {
        'EXPO_TIME': [[0] for _ in range(8)],
        'SEG_COOR_CFG': [[0 for _ in range(16)] for _ in range(8)],
        'V_SEG_EN': [[0] for _ in range(8)],
        'H_18SPAD_EN': [[0] for _ in range(8)],
        'SLOT_SHOTNUM': [[0] for _ in range(8)],
        'SLOT_IDLETIME': [[0] for _ in range(8)],
        'EXPO_TRGO_EN': [[0] for _ in range(8)],
        'SST_MODE_SEL': [[0] for _ in range(8)],
        'EXPO_LASPRD': [[0] for _ in range(8)],
        'SEG_TRGO_SEL': [[0 for _ in range(16)] for _ in range(8)],
        'EXPO_PLSWF': [[0 for _ in range(4)] for _ in range(8)],
        'EXPO_PLSWC': [[0 for _ in range(4)] for _ in range(8)],
        'TRG_DELAYH': [[0 for _ in range(4)] for _ in range(8)],
        'TDC_DELAYH': [[0 for _ in range(4)] for _ in range(8)],
        'PRBS_INTERVAL': [[0 for _ in range(4)] for _ in range(8)],
        'PRBS_STEP': [[0 for _ in range(4)] for _ in range(8)],
        'PRBS_SEL': [[0 for _ in range(4)] for _ in range(8)],
        'PRBS_EN': [[0 for _ in range(4)] for _ in range(8)],
        'PRBS_SEED': [[0 for _ in range(4)] for _ in range(8)],
        'SHOT_PXL_SEQ': [[0 for _ in range(256)] for _ in range(8)],
        'ECHO_JDG_WIDTH_THRS': [[0] for _ in range(8)],
        'ECHO_SPLIT_WIDTH_THRS': [[0] for _ in range(8)],
        'ECHO_VALL_THRS': [[0] for _ in range(8)],
        'INTF_CMP_WIN': [[0] for _ in range(8)],
        'NOS_MIRROR_EN': [[0] for _ in range(8)],
        'FS_X_SHIFT': [[0 for _ in range(16)] for _ in range(8)],
        'FS_X_VALUE': [[0 for _ in range(16)] for _ in range(8)],
        'FS_Y_MAXVAL': [[0] for _ in range(8)],
        'FS_Y_VALUE': [[0 for _ in range(16)] for _ in range(8)],
        'JDG_OFFSET': [[0] for _ in range(8)],
        'JDG_THRS_COEF': [[0] for _ in range(8)],
        'DET_MIN_VAL': [[0] for _ in range(8)],
        'FIR_KN': [[0 for _ in range(16)] for _ in range(8)],
        'ACCU_GAIN_SEG': [[0 for _ in range(16)] for _ in range(8)],
        'NOS_X_SHIFT': [[0 for _ in range(16)] for _ in range(8)],
        'NOS_X_VALUE': [[0 for _ in range(16)] for _ in range(8)],
        'NOS_Y_VALUE': [[0 for _ in range(16)] for _ in range(8)]
    }
    return roi_config_dict


def read_roi_from_excel(filename, sheet_sel, dec_or_hex=16):
    roi_config_dict = roi_config_dict_ini()

    wb = op.load_workbook(filename)

    roi_sram_data = []
    if len(wb.sheetnames) < (sheet_sel + 1):
        raise ValueError(f"Excel doesn't have {PubMethod.get_ordinal(sheet_sel + 1)} sheet...")
    sheet = wb.worksheets[sheet_sel]

    for row_value in sheet.iter_rows(values_only=True):
        __data__ = list(cell if cell is not None else "" for cell in row_value)
        roi_sram_data.append(__data__[2:])

    roi_sram_data.pop(0)
    # print(roi_sram_data)
    roi_field_st_index = 0
    for key, value in roi_config_dict.items():
        roi_field_num_in_one_group = len(roi_config_dict[key][0])
        for group_cnt in range(8):
            for roi_field_cnt in range(roi_field_num_in_one_group):
                roi_config_dict[key][group_cnt][roi_field_cnt] = int(
                    str(roi_sram_data[roi_field_st_index + roi_field_cnt][group_cnt]), dec_or_hex)
        roi_field_st_index += roi_field_num_in_one_group
        # print(key, "\t:\t", roi_config_dict[key])
    return roi_config_dict


def analysis_roi_file(roi_obj):
    roi_config_dict = roi_config_dict_ini()
    angle_grp_cnt = len(roi_obj) // 674
    for grp in range(angle_grp_cnt):
        roi_offset = 674 * grp
        for i in range(16):
            roi_config_dict['SEG_COOR_CFG'][grp][i] = (roi_obj[i + roi_offset])
        roi_offset += 16
        roi_config_dict['V_SEG_EN'][grp][0] = (roi_obj[roi_offset])
        roi_offset += 1
        roi_config_dict['H_18SPAD_EN'][grp][0] = (roi_obj[roi_offset])
        roi_offset += 1
        roi_config_dict['SLOT_SHOTNUM'][grp][0] = (roi_obj[roi_offset])
        roi_offset += 1
        roi_config_dict['SLOT_IDLETIME'][grp][0] = (roi_obj[roi_offset])
        roi_offset += 1
        roi_config_dict['EXPO_TRGO_EN'][grp][0] = (roi_obj[roi_offset])
        roi_offset += 1
        roi_config_dict['EXPO_LASPRD'][grp][0] = (roi_obj[roi_offset] & 0xFFF)
        roi_config_dict['SST_MODE_SEL'][grp][0] = ((roi_obj[roi_offset] & 0x1000) >> 12)
        roi_offset += 1
        for i in range(8):
            roi_config_dict['SEG_TRGO_SEL'][grp][i] = ((roi_obj[roi_offset] >> (i * 2)) & 0x3)
        roi_offset += 1
        for i in range(8):
            roi_config_dict['SEG_TRGO_SEL'][grp][i + 8] = ((roi_obj[roi_offset] >> (i * 2)) & 0x3)
        roi_offset += 1
        for i in range(4):
            roi_config_dict['EXPO_PLSWC'][grp][i] = ((roi_obj[roi_offset] >> 5) & 0x3F)
            roi_config_dict['EXPO_PLSWF'][grp][i] = (roi_obj[roi_offset] & 0x1F)
            roi_offset += 1
        for i in range(4):
            roi_config_dict['TDC_DELAYH'][grp][i] = ((roi_obj[roi_offset] >> 8) & 0xFF)
            roi_config_dict['TRG_DELAYH'][grp][i] = (roi_obj[roi_offset] & 0xFF)
            roi_offset += 1
        for i in range(4):
            roi_config_dict['PRBS_SEED'][grp][i] = ((roi_obj[roi_offset] >> 8) & 0x7F)
            roi_config_dict['PRBS_EN'][grp][i] = ((roi_obj[roi_offset] >> 6) & 0x1)
            roi_config_dict['PRBS_SEL'][grp][i] = ((roi_obj[roi_offset] >> 4) & 0x3)
            roi_config_dict['PRBS_STEP'][grp][i] = ((roi_obj[roi_offset] >> 2) & 0x3)
            roi_config_dict['PRBS_INTERVAL'][grp][i] = (roi_obj[roi_offset] & 0x3)
            roi_offset += 1
        for i in range(256):
            tmp = roi_obj[roi_offset] + (roi_obj[roi_offset + 1] << 16)
            roi_config_dict['SHOT_PXL_SEQ'][grp][i] = tmp
            roi_offset += 2
        roi_config_dict['ECHO_VALL_THRS'][grp][0] = (roi_obj[roi_offset])
        roi_offset += 1
        roi_config_dict['INTF_CMP_WIN'][grp][0] = ((roi_obj[roi_offset] >> 4) & 0x3)
        roi_config_dict['NOS_MIRROR_EN'][grp][0] = ((roi_obj[roi_offset] >> 3) & 0x1)
        roi_offset += 1
        roi_config_dict['FS_Y_MAXVAL'][grp][0] = ((roi_obj[roi_offset] >> 6) & 0x3FF)
        roi_config_dict['FS_X_SHIFT'][grp][0] = (roi_obj[roi_offset] & 0xF)
        roi_offset += 1
        for i in range(1, 16):
            roi_config_dict['FS_X_SHIFT'][grp][i] = (roi_obj[roi_offset])
            roi_offset += 1
        for i in range(16):
            roi_config_dict['FS_X_VALUE'][grp][i] = (roi_obj[roi_offset])
            roi_offset += 1
        for i in range(16):
            roi_config_dict['FS_Y_VALUE'][grp][i] = (roi_obj[roi_offset])
            roi_offset += 1
        roi_config_dict['JDG_OFFSET'][grp][0] = ((roi_obj[roi_offset] >> 6) & 0x3FF)
        roi_config_dict['JDG_THRS_COEF'][grp][0] = (roi_obj[roi_offset] & 0x3F)
        roi_offset += 1
        roi_config_dict['DET_MIN_VAL'][grp][0] = (roi_obj[roi_offset])
        roi_offset += 1
        for i in range(8):
            roi_config_dict['FIR_KN'][grp][2 * i] = ((roi_obj[roi_offset] >> 8) & 0xFF)
            roi_config_dict['FIR_KN'][grp][2 * i + 1] = (roi_obj[roi_offset] & 0xFF)
            roi_offset += 1
        roi_offset += 1  # reserved
        for i in range(16):
            roi_config_dict['ACCU_GAIN_SEG'][grp][i] = (roi_obj[roi_offset])
            roi_offset += 1
        for i in range(16):
            roi_config_dict['NOS_X_SHIFT'][grp][i] = (roi_obj[roi_offset])
            roi_offset += 1
        for i in range(16):
            roi_config_dict['NOS_X_VALUE'][grp][i] = (roi_obj[roi_offset])
            roi_offset += 1
        for i in range(16):
            roi_config_dict['NOS_Y_VALUE'][grp][i] = (roi_obj[roi_offset])
            roi_offset += 1


def roi_sram_generate(roi_config, group_num):
    roi_data = []
    for angle_grp_cnt in range(group_num):
        for i in range(16):
            roi_data.append((roi_config['SEG_COOR_CFG'][angle_grp_cnt][i] & 0x7F))
        roi_data.append((roi_config['V_SEG_EN'][angle_grp_cnt][0] & 0xFFFF))
        roi_data.append((roi_config['H_18SPAD_EN'][angle_grp_cnt][0] & 0x1F))
        roi_data.append((roi_config['SLOT_SHOTNUM'][angle_grp_cnt][0] & 0xFF))
        roi_data.append((roi_config['SLOT_IDLETIME'][angle_grp_cnt][0] & 0xFFFF))
        roi_data.append((roi_config['EXPO_TRGO_EN'][angle_grp_cnt][0] & 0xF))
        roi_data.append(((roi_config['SST_MODE_SEL'][angle_grp_cnt][0] & 0x1) << 12) +
                        (roi_config['EXPO_LASPRD'][angle_grp_cnt][0] & 0xFFF))
        for j in range(2):
            data = 0
            for i in range(8):
                data += ((roi_config['SEG_TRGO_SEL'][angle_grp_cnt][i + j * 8] & 0x3) << (i * 2))
            roi_data.append(data)

        for j in range(4):
            roi_data.append(((roi_config['EXPO_PLSWC'][angle_grp_cnt][j] & 0x3F) << 5) +
                            (roi_config['EXPO_PLSWF'][angle_grp_cnt][j] & 0x1F))
            roi_data.append(((roi_config['TDC_DELAYH'][angle_grp_cnt][j] & 0xFF) << 8) +
                            (roi_config['TRG_DELAYH'][angle_grp_cnt][j] & 0xFF))
            roi_data.append(((roi_config['PRBS_SEED'][angle_grp_cnt][j] & 0x7F) << 8) +
                            ((roi_config['PRBS_EN'][angle_grp_cnt][j] & 0x1) << 6) +
                            ((roi_config['PRBS_SEL'][angle_grp_cnt][j] & 0x3) << 4) +
                            ((roi_config['PRBS_STEP'][angle_grp_cnt][j] & 0x3) << 2) +
                            (roi_config['PRBS_INTERVAL'][angle_grp_cnt][j] & 0x3))
        for i in range(256):
            roi_data.append((roi_config['SHOT_PXL_SEQ'][angle_grp_cnt][i] & 0xFFFF))
            roi_data.append(((roi_config['SHOT_PXL_SEQ'][angle_grp_cnt][i] >> 16) & 0xFFFF))
        roi_data.append(((roi_config['ECHO_JDG_WIDTH_THRS'][angle_grp_cnt][0] & 0xFF) << 8) +
                        (roi_config['ECHO_SPLIT_WIDTH_THRS'][angle_grp_cnt][0] & 0xFF))
        roi_data.append((roi_config['ECHO_VALL_THRS'][angle_grp_cnt][0] & 0x7))
        roi_data.append(((roi_config['INTF_CMP_WIN'][angle_grp_cnt][0] & 0x3) << 4) +
                        ((roi_config['NOS_MIRROR_EN'][angle_grp_cnt][0] & 0x1) << 3))
        roi_data.append(((roi_config['FS_Y_MAXVAL'][angle_grp_cnt][0] & 0x3FF) << 6) +
                        (roi_config['FS_X_SHIFT'][angle_grp_cnt][0] & 0xF))
        for i in range(1, 16):
            roi_data.append((roi_config['FS_X_SHIFT'][angle_grp_cnt][i] & 0xF))
        for i in range(16):
            roi_data.append((roi_config['FS_X_VALUE'][angle_grp_cnt][i] & 0xFFF))
        for i in range(16):
            roi_data.append((roi_config['FS_Y_VALUE'][angle_grp_cnt][i] & 0x3FF))
        roi_data.append(((roi_config['JDG_OFFSET'][angle_grp_cnt][0] & 0x3FF) << 6) +
                        (roi_config['JDG_THRS_COEF'][angle_grp_cnt][0] & 0x3F))
        roi_data.append((roi_config['DET_MIN_VAL'][angle_grp_cnt][0] & 0x3FF))
        for i in range(8):
            roi_data.append(((roi_config['FIR_KN'][angle_grp_cnt][2 * i + 1] & 0xFF) << 8) +
                            (roi_config['FIR_KN'][angle_grp_cnt][2 * i] & 0xFF))
        roi_data.append(0)  # kn_sum, Rev
        for i in range(16):
            roi_data.append((roi_config['ACCU_GAIN_SEG'][angle_grp_cnt][i] & 0xFFF))
        for i in range(16):
            roi_data.append((roi_config['NOS_X_SHIFT'][angle_grp_cnt][i] & 0x7))
        for i in range(16):
            roi_data.append((roi_config['NOS_X_VALUE'][angle_grp_cnt][i] & 0x3FF))
        for i in range(16):
            roi_data.append((roi_config['NOS_Y_VALUE'][angle_grp_cnt][i] & 0xFF))
    return roi_data


def get_prbs_result(n_prbs: int, seed: int, sel: int, step: int) -> int:
    """
    单次SLOT的PRBS delay值计算，返回PRBS delay的总和

    Args:
        n_prbs: 插入PRBS次数
        seed: PRBS种子
        sel: PRBS类型
        step: PRBS倍数

    Returns: 单次SLOT的PRBS delay的总和

    """
    prbs_result = []
    prbs = seed
    prbs_sel = sel
    prbs_step = 2 ** step
    # prbs_intv = 2 ** interval

    for i in range(n_prbs):
        if prbs_sel == 0:
            xor_bit = ((prbs >> 3) & 0x1) ^ ((prbs >> 2) & 0x1)
            result = ((prbs << 1) | xor_bit) & 0xF
            prbs_result.append(result * prbs_step)
        elif prbs_sel == 1:
            xor_bit = ((prbs >> 4) & 0x1) ^ ((prbs >> 3) & 0x1)
            result = ((prbs << 1) | xor_bit) & 0x1F
            prbs_result.append(result * prbs_step)
        elif prbs_sel == 2:
            xor_bit = ((prbs >> 5) & 0x1) ^ ((prbs >> 4) & 0x1)
            result = ((prbs << 1) | xor_bit) & 0x3F
            prbs_result.append(result * prbs_step)
        else:  # if prbs_sel == 3:
            xor_bit = ((prbs >> 6) & 0x1) ^ ((prbs >> 5) & 0x1)
            result = ((prbs << 1) | xor_bit) & 0x7F
            prbs_result.append(result * prbs_step)
        prbs = result
    prbs_sum = sum(prbs_result)
    return prbs_sum


def expo_time_cal(csru_cfg: dict, roi_config: dict, grp_sel: int) -> int:
    """
    SWAN:
    单次ANGLE_GRP的曝光时间计算，输入对应ANGLE_GRP的ROI SRAM配置，计算得出曝光时间

    Args:
        csru_cfg:CSRU寄存器输入
        roi_config: ANGLE_GRP ROI寄存器输入
        grp_sel: 当前 group_index

    Returns: 单次ANGLE_GRP的曝光时间, unit: cyc

    """
    ulr_en = csru_cfg["ULR_EN"]
    hop_en = csru_cfg["LSPRD_HOP_EN"]
    hop_cnt = csru_cfg["LSPRD_HOP_CNTS"] + 1
    hop_step = csru_cfg["LSPRD_HOP_STEP"]
    hist_maxbin_thrs = csru_cfg["HIST_MAXBIN_THRS"]

    sub_shotnum = roi_config["SLOT_SHOTNUM"][grp_sel][0]
    expo_lasprd = roi_config["EXPO_LASPRD"][grp_sel][0]
    trgo_en = roi_config["EXPO_TRGO_EN"][grp_sel][0]
    prbs_seed = roi_config["PRBS_SEED"][grp_sel]
    prbs_sel = roi_config["PRBS_SEL"][grp_sel]
    prbs_step = roi_config["PRBS_STEP"][grp_sel]
    prbs_interval = roi_config["PRBS_INTERVAL"][grp_sel]
    prbs_en = roi_config["PRBS_EN"][grp_sel]
    tdc_delayh = roi_config["TDC_DELAYH"][grp_sel]

    trig_channel = 4
    prbs_inject = []
    expo_time_pre = 0
    delay_last = []
    prbs_sum = []
    # TDC_SYNC拼接的计算
    if ulr_en == 3:
        act_expo_lasprd = expo_lasprd * 4
    else:
        act_expo_lasprd = expo_lasprd

    # 跳频时间的计算
    for shot_cnt in range(sub_shotnum - 1):
        hop_add_cnt = int(shot_cnt / hop_cnt)
        if hop_en and hop_add_cnt > 0:
            hop_expo_time = hop_step * hop_add_cnt
        else:
            hop_expo_time = 0
        expo_time_pre = expo_time_pre + hop_expo_time + act_expo_lasprd + 1
    # PRBS时间的计算
    for chnl in range(trig_channel):
        prbs_cnt_sum = int((sub_shotnum - 1) / (2 ** prbs_interval[chnl]))
        if prbs_en and ((sub_shotnum - 1) % (2 ** prbs_interval[chnl]) == 0):
            prbs_inject.append(1)
            prbs_rs = get_prbs_result(prbs_cnt_sum, prbs_seed[chnl], prbs_sel[chnl], prbs_step[chnl])
            prbs_sum.append(prbs_rs)
        else:
            prbs_inject.append(0)
            prbs_sum.append(0)
    # TRIGO使能的Gating
    for chnl in range(trig_channel):
        if (trgo_en >> chnl) & 0x1:
            if prbs_inject[chnl]:
                delay_last.append(tdc_delayh[chnl] + prbs_sum[chnl])
            else:
                delay_last.append(tdc_delayh[chnl])
        else:
            delay_last.append(0)
    # 不同通道的最大曝光时间输出
    tdc_dly_max = 0 if trgo_en == 0 else max(delay_last)
    expo_time_cyc = tdc_dly_max + expo_time_pre + (hist_maxbin_thrs+1)*2

    return expo_time_cyc


if __name__ == '__main__':
    pass

