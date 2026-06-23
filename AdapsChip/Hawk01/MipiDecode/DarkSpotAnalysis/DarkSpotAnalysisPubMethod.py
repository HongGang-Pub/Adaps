import numpy as np
# import openpyxl
# from openpyxl.styles import PatternFill, Border, Side


def data_process(array_lst, coefficient_list, ref_array_idx=0, ref_array_lst=None):
    """
    查找黑点: 对数据进行处理，按照要求过滤符合条件的点
    Args:
        ref_array_lst ():
        array_lst ():
        coefficient_list ():
        ref_array_idx ():

    Returns:
        tuple: spad_bin_number_dict: 当前当前array_lst过滤点数据信息； base_num_list: 基准值
    """

    frame_number = len(array_lst)
    # image_idx = list(index for index in range(frame_number))
    base_num_list = []

    # __array__ = np.zeros((2, 3))
    # base_num = __array__.max()
    err_coor_record_dict = {}
    spad_bin_number_dict = {}
    coefficient_list.sort(reverse=True)
    for coefficient in coefficient_list:
        err_coor_record_dict[coefficient] = []
        spad_bin_number_dict[coefficient] = []

    for index in range(frame_number):
        _array = array_lst[index]
        base_num = np.median(_array)

        if ref_array_lst is not None:
            _ref_array = ref_array_lst[index]
            base_num = np.median(_ref_array)
        base_num_list.append(base_num)

        if index < ref_array_idx:
            continue

        # image_idx[index] = "{}: {}".format(index, base_num)

        array_spec = _array.shape
        rows = array_spec[0]
        cols = array_spec[1]
        # print("rows: {} , cols: {}".format(rows, cols))
        for c in range(cols):
            for r in range(rows):
                value = _array[r, c]
                if value == 0:
                    continue
                for coefficient in coefficient_list:
                    if value < base_num * coefficient:
                        try:
                            idx = err_coor_record_dict[coefficient].index((r, c))
                            spad_bin_number_dict[coefficient][idx][index + 2] = [1, value]
                            # spad_bin_number_dict[coefficient][idx][index + 2] = value
                        except:
                            # print("Add Coor [{}, {}]".format(r, c))
                            err_coor_record_dict[coefficient].append((r, c))
                            ini_list = list(-1 for i in range(frame_number + 2))
                            ini_list[0:2] = [r, c]
                            ini_list[index + 2] = [1, value]
                            # ini_list[index + 2] = value
                            spad_bin_number_dict[coefficient].append(ini_list)
                    else:
                        break

    # print(spad_array_analysis)
    # 对没有问题的点 进行 bin_number 的填充
    for coefficient in coefficient_list:
        for point_cnt in range(len(spad_bin_number_dict[coefficient])):
            [x, y] = spad_bin_number_dict[coefficient][point_cnt][0:2]
            bin_number_cnt = spad_bin_number_dict[coefficient][point_cnt][2:]
            for i in range(frame_number):
                symbols = 1 if i < ref_array_idx else 0
                if bin_number_cnt[i] == -1:
                    spad_bin_number_dict[coefficient][point_cnt][i + 2] = [symbols, array_lst[i][x, y]]
                    # spad_bin_number_dict[coefficient][point_cnt][index + 2] = array_lst[index][x, y]

    # title = ["x", "y"] + image_idx
    # title = ["x", "y"] + ["D2"]
    #
    # if fd_path != "None":
    #     fname = "Spad分析"
    #     fd_path = "Array_List"
    #     for coefficient in coefficient_list:
    #         sheet_name = "{}_{}".format(chip_numbers, coefficient)
    #         spad_bin_number_dict[coefficient].insert(0, title)
    #         save_excel(fname, sheet_name, spad_bin_number_dict[coefficient], fd_path)
    return spad_bin_number_dict, base_num_list


# custom_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                        top=Side(style='thin'), bottom=Side(style='thin'))


def write_excel(data, excel_name):
    """Edit by songlin.yin: 将数据写入Excel"""

    import openpyxl
    from openpyxl.styles import PatternFill, Border, Side
    custom_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    # 创建一个新的 Excel 工作簿
    workbook = openpyxl.Workbook()
    workbook.remove(workbook["Sheet"])

    for chip_number in data["chip_numbers"]:
        sorted_indices = [index for index, value in
                          sorted(enumerate(data["axis1"][chip_number]), key=lambda x: x[1])]
        for coefficient in data["coefficients"]:
            sheet = workbook.create_sheet(title=f"{chip_number}_{coefficient}")
            reg_list = ["{}:{}".format(data["axis1"][chip_number][index], data["base_number"][chip_number][index]) for
                        index in sorted_indices]
            reg_list.insert(0, "y")
            reg_list.insert(0, "x")
            sheet.append(reg_list)
            row = 0
            for inner_list in data[chip_number][coefficient]:
                sheet.append([inner_list[0], inner_list[1]])
                col = 0
                flag = ""
                for axis1 in sorted_indices:
                    if data[chip_number][coefficient][row][2 + axis1][0]:
                        fille = PatternFill('solid', fgColor='FFFFFF')  # 设置填充颜色为白色
                        # fille = GradientFill(stop=("2F0000", "EF0000"))
                    else:
                        fille = PatternFill('solid', fgColor='B0B0B0')  # 设置填充颜色为灰色

                    post_count = data[chip_number][coefficient][row][2 + axis1][1]
                    if col > 0:
                        count = data[chip_number][coefficient][row][2 + pre_axis1][1]
                        if post_count > count or (count - post_count) < 50:
                            flag = flag + "↗"  # ↗ ↑
                        else:
                            flag = flag + "↘"  # ↘ ↓

                    pre_axis1 = axis1
                    sheet.cell(row=row + 2, column=col + 3, value=post_count).fill = fille
                    col = col + 1
                sheet.cell(row=row + 2, column=col + 3, value=flag)

                row = row + 1
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = custom_border

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("保存失败:", e)
    finally:
        # 无论是否保存成功，都关闭 Excel 文件
        workbook.close()


def cmp_data(coefficient, sv11, data, excel_name):
    """Edit by songlin.yin: 比较模组之间是否又重复的点"""

    import openpyxl
    from openpyxl.styles import PatternFill, Border, Side
    custom_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

    m_list = []
    for chip_number in data["axis1"]:
        if sv11 in data["axis1"][chip_number]:
            m_list.append(chip_number)
    all_coor_list = []
    for chip_number in m_list:
        coor_list = []
        index = data["axis1"][chip_number].index(sv11)
        for n in data[chip_number][coefficient]:
            if n[2 + index][0] == 1:
                coor_list.append([n[0], n[1]])
        all_coor_list.append(coor_list)
    write_list = [element for sublist in all_coor_list for element in sublist]
    # res = list(set(map(lambda index: tuple(sorted(index)), write_list)))
    # write_list = [list(item) for item in res]

    tmp_list = []
    for coor in write_list:
        if not coor in tmp_list:
            tmp_list.append(coor)

    write_list = tmp_list
    for chip_number in write_list:
        pb = []
        for j in all_coor_list:
            if chip_number in j:
                pb.append("√")
            else:
                pb.append("×")
        index = write_list.index(chip_number)
        write_list[index] = chip_number + pb

    m_list.insert(0, 'y')
    m_list.insert(0, 'x')
    wk = openpyxl.load_workbook(excel_name)
    sheet = wk.create_sheet(title=f"cp_{coefficient}_{sv11}")
    sheet.append(m_list)
    for chip_number in write_list:
        sheet.append(chip_number)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = custom_border
            if cell.value == "×":
                cell.fill = PatternFill('solid', fgColor='B0B0B0')

    try:
        wk.save(excel_name)
    except Exception as chip_number:
        print("保存失败:", chip_number)
    finally:
        # 无论是否保存成功，都关闭 Excel 文件
        wk.close()
