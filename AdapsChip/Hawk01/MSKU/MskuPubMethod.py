import csv
import logging
import re
import os
import openpyxl

import numpy as np
import matplotlib.animation as animation
import matplotlib.pyplot as plt

from SelfDefinedPackge import PubMethod
from SelfDefinedPackge import ArrayPubMethod
from matplotlib.pyplot import MultipleLocator


def zone_config(sub_expotime, sub_idletime, expo_lasprd, expo_plswc, expo_plswf, tx_en, spad_en_in3rows, kernel):
    """ 根据传入参数生成每个分区的 Zone Config """
    if not (isinstance(kernel, list)):
        raise ValueError("Kernel is no list!")
    if len(kernel) != 8:
        raise ValueError("Kernel size error!")

    zone_cfg = [sub_expotime, sub_idletime, expo_lasprd, (int(expo_plswc << 5) + expo_plswf),
                (int(tx_en << 2) + spad_en_in3rows)]
    zone_cfg = zone_cfg + kernel
    return zone_cfg


def ZonesConfigGenerate(cfg):
    """ 调用zon_config方法生成所有分区的 Zone Config """
    kernel = []
    zone_mem = []

    for v_roll_cnt in range(0, cfg['V_ROLL_NUM'] + 1):
        if cfg['zone_cfg_sel'] != -1:
            zone_sel = "Zone{}".format(cfg['zone_cfg_sel'])
            kernel_sel = "Zone_{}_MF_KN".format(cfg['zone_cfg_sel'])
        else:
            zone_sel = "Zone{}".format(v_roll_cnt)
            kernel_sel = "Zone_{}_MF_KN".format(v_roll_cnt)

        if len(cfg['zone_cfg_def'][kernel_sel]) != 16:
            raise ValueError("Kernel:{} config Error!".format(kernel_sel))

        for i in range(0, 8):
            kernel.append(
                int(cfg['zone_cfg_def'][kernel_sel][i * 2 + 1] << 8) + cfg['zone_cfg_def'][kernel_sel][i * 2])

        try:
            per_zone_mem = zone_config(sub_expotime=cfg['zone_cfg_def']['SUB_EXPOTIME'][zone_sel],
                                       sub_idletime=cfg['zone_cfg_def']['SUB_IDLETIME'][zone_sel],
                                       expo_lasprd=cfg['zone_cfg_def']['EXPO_LASPRD'][zone_sel],
                                       expo_plswc=cfg['zone_cfg_def']['EXPO_PLSWC'][zone_sel],
                                       expo_plswf=cfg['zone_cfg_def']['EXPO_PLSWF'][zone_sel],
                                       tx_en=cfg['zone_cfg_def']['TX_EN'][zone_sel],
                                       spad_en_in3rows=cfg['zone_cfg_def']['SPADEN_IN3ROWS'][zone_sel],
                                       kernel=kernel)
            kernel = []
        except BaseException as msg:
            raise ValueError("{} configuration may be missing or incorrect! Log: {}".format(zone_sel, msg))
        zone_mem.append(per_zone_mem)
    return zone_mem


def do_mark(info, fontsize=5):
    (x, y, text) = info
    plt.text(x + 5, y + 3, text, fontdict={
        'family': 'Times New Roman',  # 标注文本字体
        'fontsize': fontsize,  # 文本大小
        'fontweight': 'bold',  # 字体粗细
        # 'fontstyle': 'italic',  # 字体风格
        'color': 'white',  # 文本颜色
        'backgroundcolor': 'blue',  # 背景颜色
        'bbox': {
            'boxstyle': 'round',  # 椭圆外框
            'edgecolor': 'white',  # 线框颜色
            'linewidth': 0
        }
    })

    # plt.annotate(s, xy=(x, y), xytext=(x-30, y+3),
    #              arrowprops={
    #                  'headwidth': 12,  # 箭头头部的宽度
    #                  'headlength': 8,  # 箭头头部的长度
    #                  'width': 3,  # 箭头尾部的宽度
    #                  'facecolor': 'w',  # 箭头的颜色
    #                  'shrink': 0.1,  # 从箭尾到标注文本内容开始两端空隙长度
    #              },
    #              family='Times New Roman',  # 标注文本字体为Times New Roman
    #              fontsize=10,  # 文本大小为18
    #              fontweight='bold',  # 文本为粗体
    #              color='white',  # 文本颜色为红色
    #              backgroundcolor='black'
    #              # ha = 'center' # 水平居中
    #              )


def ParseRoiMem(hawk01_config, roi_file=None, f_path=None):
    """
    根据寄存器配置解析 ROI 数据
    Args:
        hawk01_config (寄存器配置相关信息):
        roi_file (str): 指定 ROI 文件路径，为 None 时，从cfg中获取 roi_file
        f_path (str): 解析ROI后 masking 成图效果展示

    Returns:
        tuple: zone_roi_mem, msku_roi_mem
    """
    scan_mode = hawk01_config["SCAN_MODE"]
    v_roll_num = hawk01_config["V_ROLL_NUM"]
    h_roll_num = hawk01_config["H_ROLL_NUM"]
    h_vld_seg = hawk01_config["H_VLD_SEG"]

    if roi_file is None:
        roi_file = hawk01_config["roi_file"]
    __roi_data__ = PubMethod.read_file(roi_file)

    roi_mem = []
    zone_roi_mem = []
    msku_roi_mem = []

    if len(__roi_data__[0].strip()) == 2:   # Byte
        for r in range(0, len(__roi_data__)//2):
            try:
                roi_data = int(__roi_data__[r*2], 16) + int(__roi_data__[r*2+1], 16) * 256
                roi_mem.append(roi_data)
            except BaseException as e:
                raise ValueError(f"ROI format error:{e}")
    elif len(__roi_data__[0].strip()) == 4:     # Half-word
        for r in range(0, len(__roi_data__)):
            try:
                roi_data = int(__roi_data__[r], 16)
                roi_mem.append(roi_data)
            except BaseException as e:
                raise ValueError(f"ROI format error:{e}")
    else:
        raise ValueError("ROI format error...")

    roi_len = h_vld_seg + 1 if scan_mode == 0 else h_roll_num + 1
    zone_len = roi_len * 6 + 13

    if zone_len * (v_roll_num + 1) != len(roi_mem):
        raise ValueError("ROI data does not match the register configuration, "
                         "Please check ROI or SCAN_MODE, V_ROLL_NUM, H_ROLL_NUM, H_VLD_SEG!!!")

    for v_roll_cnt in range(v_roll_num + 1):
        seg_coor_st_index = v_roll_cnt * zone_len + 13

        zone_mem = roi_mem[seg_coor_st_index - 13:seg_coor_st_index]
        zone_roi_mem.append(zone_mem)

        msku_mem = roi_mem[seg_coor_st_index: seg_coor_st_index + zone_len - 13]
        msku_roi_mem.append(msku_mem)

    # 成图展示 masking 效果
    if f_path is not None:
        RollingArrayCollect(msku_roi_data=msku_roi_mem, cfg=hawk01_config, is_save=1, fd_path=f_path)

    return zone_roi_mem, msku_roi_mem


def RollingArrayCollect(msku_roi_data, cfg, is_save=0, fd_path='.') -> tuple:
    """
    对 rolling 的数据生成二维数组用于成图, 包含:
        1. 单次rolling masking的二维数组;
        2. 所有rolling masking的叠加的二维数组(支持保存);
        3. 所有rolling masking叠加的深度二维数组(支持保存);
    """
    masking_array = np.zeros((576, 768), dtype=np.float32)

    scan_mode = cfg["SCAN_MODE"]
    v_roll_num = cfg['V_ROLL_NUM']
    h_roll_num = cfg['H_ROLL_NUM']
    h_vld_seg = cfg['H_VLD_SEG']
    masking_coor_info = []

    masking_arrays = []
    pcm_array = np.zeros((576, 768), dtype=np.float32)
    ptm_array = np.zeros((192, 256), dtype=np.float32)

    if scan_mode == 0:
        for vroll_cnt in range(v_roll_num + 1):
            per_rolling_data = msku_roi_data[vroll_cnt]
            dsp = (vroll_cnt * 2) % 32 + 10
            # for per_coor in per_rolling_data:
            for seg_cnt in range(len(per_rolling_data)):
                per_coor = per_rolling_data[seg_cnt]
                seg_num = per_coor >> 10
                spad_coor = per_coor % 1024
                # spad_array[spad_coor:spad_coor + 3, seg_num * 48:(seg_num + 1) * 48, :] = np.arrays(
                #     [dsp, dsp, dsp])
                masking_array[spad_coor:spad_coor + 3, seg_num * 48:(seg_num + 1) * 48] = dsp
                pcm_array[spad_coor:spad_coor + 3, seg_num * 48:(seg_num + 1) * 48] = dsp
                try:
                    ptm_array[spad_coor // 3, seg_num * 16:(seg_num + 1) * 16] = dsp
                except:
                    pass
                if seg_cnt == 0:
                    masking_coor_info.append([seg_num * 48, spad_coor, "1D VROll_{}".format(vroll_cnt+1)])
            masking_arrays.append(masking_array)
            masking_array = np.zeros((576, 768))
    else:
        roll_cnt = 0
        for vroll_cnt in range(v_roll_num + 1):
            per_zone_data = msku_roi_data[vroll_cnt]
            for hroll_cnt in range(h_roll_num + 1):
                dsp = (vroll_cnt*(h_roll_num+1) + hroll_cnt) % ((h_roll_num + 1)*2) + 10
                index = hroll_cnt * 6
                per_rolling_data = per_zone_data[index: index + 6]
                # for per_coor in per_rolling_data:
                for seg_cnt in range(len(per_rolling_data)):
                    per_coor = per_rolling_data[seg_cnt]
                    seg_num = per_coor >> 10
                    spad_coor = per_coor % 1024
                    masking_array[spad_coor:spad_coor + 3, seg_num * 48:(seg_num + h_vld_seg + 1) * 48] = dsp
                    pcm_array[spad_coor:spad_coor + 3, seg_num * 48:(seg_num + h_vld_seg + 1) * 48] = dsp
                    try:
                        ptm_array[spad_coor // 3, seg_num * 16:(seg_num + h_vld_seg + 1) * 16] = dsp
                    except:
                        pass
                    if seg_cnt == 0:
                        masking_coor_info.append([seg_num * 48, spad_coor, "2D ROll_{}_{}".format(vroll_cnt+1, hroll_cnt+1)])
                roll_cnt += 1
                masking_arrays.append(masking_array)
                masking_array = np.zeros((576, 768))
    if is_save:
        # 对 acc_spad_array 和 depth_spad_array进行保存
        fig = plt.figure(clear=True)
        ax = fig.gca()
        ax.xaxis.tick_top()  # 设置x坐标轴位置在顶部
        ax.yaxis.set_major_locator(MultipleLocator(50))
        ax.xaxis.set_major_locator(MultipleLocator(48))
        # ax.imshow(spad_array, cmap="gray")
        ax.imshow(pcm_array)
        # plt.show()
        for info in masking_coor_info:
            do_mark(info)
        ArrayPubMethod.ArrayImageSave(fname='imag_msku', fd_path=fd_path)
        plt.close()

        fig = plt.figure(clear=True)
        ax = fig.gca()
        ax.xaxis.tick_top()  # 设置x坐标轴位置在顶部
        ax.yaxis.set_major_locator(MultipleLocator(20))
        ax.xaxis.set_major_locator(MultipleLocator(16))
        ax.imshow(ptm_array)
        # plt.show()
        ArrayPubMethod.ArrayImageSave(fname="imag_depth", fd_path=fd_path)
        plt.close()

        # plt.show()
        # plt.imsave("{}/msku_imag.png".format(fd_path), spad_array, dpi=600)
    return masking_arrays, pcm_array, ptm_array, masking_coor_info


# def animation_img(fig, msku_roi_data, cfg):
#     # ax = plt.gca()
#     # fig = plt.figure()
#     arrays, acc_spad_array, depth_spad_array, info = RollingArrayCollect(msku_roi_data, cfg)
#
#     fig.clf()
#     ax = fig.add_subplot(111)
#
#     # --------------------- 配置刻度 --------------------
#     x_major_locator = MultipleLocator(48)
#     # 把x轴的刻度间隔设置为 48，并存在变量里
#     y_major_locator = MultipleLocator(50)
#     # 把y轴的刻度间隔设置为 50，并存在变量里
#
#     # ax为两条坐标轴的实例
#     ax.xaxis.set_major_locator(x_major_locator)
#     ax.yaxis.set_major_locator(y_major_locator)
#
#     # --------------------- 配置坐标轴属性刻度 --------------------
#     ax.xaxis.tick_top()     # 设置x坐标轴位置在顶部
#     ax.spines['top'].set_color('gray')
#     ax.spines['top'].set_linewidth(5)
#     ax.spines['left'].set_color('gray')
#     ax.spines['left'].set_linewidth(5)
#     # 将侧轴、顶部轴设置为None
#     ax.spines['right'].set_color(None)
#     ax.spines['bottom'].set_color(None)
#
#     # im = ax.imshow(X=arrays[0], cmap="gray")
#
#     # def init():
#     #     im.set_data(arrays[0])
#     #     return im
#
#     def update(i):
#         # im.set_data(arrays[index])
#         imgs = ax.imshow(X=arrays[i], cmap="gray")
#
#         x, y, s = info[i]
#         title = ax.text(x + 5, y + 3, s, fontdict={
#             'family': 'Times New Roman',  # 标注文本字体
#             'fontsize': 10,  # 文本大小
#             'fontweight': 'bold',  # 字体粗细
#             # 'fontstyle': 'italic',  # 字体风格
#             'color': 'white',  # 文本颜色
#             'backgroundcolor': 'blue',  # 背景颜色
#             'bbox': {
#                 'boxstyle': 'round',  # 椭圆外框
#                 'edgecolor': 'white',  # 线框颜色
#                 'linewidth': 0
#             }
#         })
#         return [imgs] + [title]
#
#     ani = animation.FuncAnimation(fig, update, range(len(arrays)), interval=700, blit=True)
#
#     # plt.show()
#     # plt.close()
#     return ani


def DirectAccessCaliDataByTXT(hawk01_config):
    """通过读取文件的形式获取 cali_data"""
    ini_cali_datas = PubMethod.read_file(hawk01_config["cali_file"])

    # 去除单行注释
    # /////////////////////////////////////////////////////////////////////
    cali_datas = []
    for line_cnt in range(len(ini_cali_datas)):
        _str = ini_cali_datas[line_cnt]
        if _str.strip() == '\n':
            continue
        elif _str.strip()[0:2] == '//':
            continue
        else:
            cali_datas.append([_str, line_cnt + 1])

    # 校验标定数量是否正确
    # /////////////////////////////////////////////////////////////////////
    num = (hawk01_config['V_ROLL_NUM'] + 1) * (hawk01_config['H_ROLL_NUM'] + 1) if hawk01_config['SCAN_MODE'] == 1 else (hawk01_config['V_ROLL_NUM'] + 1)
    if len(cali_datas) < num:  # 标定数量少于配置所需标定数时, 结束程序
        info = (f"Preview failed! Log: Based on the configuration information of V_ROLL_NUM & H_ROLL_NUM, "
                f"{num} cali data are required, but only {len(cali_datas)} cali data are available.")
        raise ValueError(info)
    elif len(cali_datas) > num:  # 标定数量多余所需标定数时, 打印提示信息, 提示配置信息与标定信息不匹配
        logging.warning(f"Be careful! The calibration data may not match the register configuration.")

    def _split_cali_data(index):
        [data, lines] = cali_datas[index]
        data = re.split(',|;|，|；|//', data)
        # if len(data) < 2:
        #     raise ValueError(f"Calibration data format error.\n"
        #                      f"line{lines}: {data}")
        try:
            _start_index = int(data[1])
            _seg_num = int(data[0]) // 48
        except BaseException as e:
            raise ValueError(f"Calibration data format error.\n"
                             f"line{lines}: {data}")
        if _seg_num > 15:
            raise ValueError(f"Calibration data error.\n"
                             f"line{lines}: {data}\n"
                             f"Error: {data[0]} beyond 767.")
        return _seg_num, _start_index

    img_roi_data = []
    per_img_roi_data = []  # 存储一张PCM灰度图获取的ROI数据

    frame_cnt = 0
    if hawk01_config['SCAN_MODE'] == 0:
        for vroll_cnt in range(0, hawk01_config['V_ROLL_NUM'] + 1):
            seg_num, start_index = _split_cali_data(frame_cnt)

            for seg_cnt in range(0, hawk01_config['H_VLD_SEG'] + 1):
                per_img_roi_data.append([seg_num + seg_cnt, start_index])

            img_roi_data.append(per_img_roi_data)
            per_img_roi_data = []
            frame_cnt += 1
    else:
        for vroll_cnt in range(0, hawk01_config['V_ROLL_NUM'] + 1):
            for hroll_cnt in range(0, hawk01_config['H_ROLL_NUM'] + 1):
                seg_num, start_index = _split_cali_data(frame_cnt)
                per_img_roi_data.append([seg_num, start_index])

                img_roi_data.append(per_img_roi_data)
                per_img_roi_data = []
                frame_cnt += 1
    return img_roi_data


def DirectAccessCaliDataByExcel(hawk01_config):
    """通过读取文件的形式获取 cali_data"""
    file = hawk01_config["cali_file"]
    sheet_sel = hawk01_config["sheet_sel"]

    file_name, file_ext = os.path.splitext(file)
    cali_datas = []
    if file_ext == ".csv":
        with open(file, newline='', encoding="utf-8") as f:
            datas = csv.reader(f, delimiter=',', quotechar='|')
            for __data__ in datas:
                cali_datas.append(__data__[1:])
    elif file_ext in [".xlsx", ".xls"]:
        # wb = xlrd.open_workbook(file)
        # if len(wb.sheet_names()) < (sheet_sel+1):
        #     raise ValueError(f"Excel doesn't have {PubMethod.get_ordinal(sheet_sel+1)} sheet...")
        # sheet = wb.sheet_by_index(sheet_sel)
        # nrows = sheet.nrows
        # for row in range(nrows):
        #     __data__ = sheet.row_values(row)
        #     cali_datas.append(__data__[1:])
            
        # Used openpyxl to read the excel file, xlrd is not support .xlsx file
        # https://openpyxl.readthedocs.io/en/stable/changes.html#xlrd-support
        wb = openpyxl.load_workbook(file)

        if len(wb.sheetnames) < (sheet_sel+1):
            raise ValueError(f"Excel doesn't have {PubMethod.get_ordinal(sheet_sel+1)} sheet...")
        sheet = wb.worksheets[sheet_sel]

        for row_value in sheet.iter_rows(values_only=True):
            __data__ = list(cell if cell is not None else "" for cell in row_value)
            cali_datas.append(__data__[1:])
    else:
        return

    cali_datas.pop(0)   # 删除第一行
    # 校验标定数量是否正确
    # /////////////////////////////////////////////////////////////////////
    num = (hawk01_config['V_ROLL_NUM'] + 1) * (hawk01_config['H_ROLL_NUM'] + 1) if hawk01_config['SCAN_MODE'] == 1 else (hawk01_config['V_ROLL_NUM'] + 1)
    if (len(cali_datas)) < num:  # 标定数量少于配置所需标定数时, 结束程序, 第一行为 Segment 标识数据
        info = (f"Preview failed! Log: Based on the configuration information of V_ROLL_NUM & H_ROLL_NUM, "
                f"{num} cali data are required, but only {len(cali_datas)} cali data are available.")
        raise ValueError(info)
    elif (len(cali_datas)) > num:  # 标定数量多余所需标定数时, 打印提示信息, 提示配置信息与标定信息不匹配
        logging.warning(f"Be careful! The calibration data may not match the register configuration.")

    img_roi_data = []
    per_img_roi_data = []  # 存储一张PCM灰度图获取的ROI数据

    frame_cnt = 0
    start_index = 0
    if hawk01_config['SCAN_MODE'] == 0:
        for vroll_cnt in range(0, hawk01_config['V_ROLL_NUM'] + 1):
            seg_hs = -1
            per_img_cali_data = cali_datas[frame_cnt]
            for i in range(16):     # 找到第一个非 0 数据, 作为 SEG_HS
                if per_img_cali_data[i] != "":
                    seg_hs = i
                    break
            if seg_hs == -1:
                raise ValueError(f"The {PubMethod.get_ordinal(frame_cnt+1)} rolling no calibration data, "
                                 f"Please fill in and try again...")
            if seg_hs + hawk01_config['H_VLD_SEG'] > 15:
                raise ValueError(f"The {PubMethod.get_ordinal(frame_cnt+1)} rolling start with {seg_hs} segment, "
                                 f"and depending on the H_VLD_SEG configuration, ROI beyond the SPAD_ARRAY.")
            for seg_cnt in range(0, hawk01_config['H_VLD_SEG'] + 1):
                seg_num = seg_hs + seg_cnt
                if per_img_cali_data[seg_num] != "":
                    try:
                        start_index = int(per_img_cali_data[seg_num])
                    except BaseException as e:
                        raise ValueError(f"The calibration data [{per_img_cali_data[seg_num]}] formatting error....")
                per_img_roi_data.append([seg_num, start_index])

            img_roi_data.append(per_img_roi_data)
            per_img_roi_data = []
            frame_cnt += 1
    else:
        for vroll_cnt in range(0, hawk01_config['V_ROLL_NUM'] + 1):
            for hroll_cnt in range(0, hawk01_config['H_ROLL_NUM'] + 1):
                seg_hs = -1
                per_img_cali_data = cali_datas[frame_cnt]
                for i in range(16):     # 找到第一个非 0 数据, 作为 SEG_HS
                    if per_img_cali_data[i] != "":
                        seg_hs = i
                        break
                if seg_hs == -1:
                    raise ValueError(f"The {PubMethod.get_ordinal(frame_cnt+1)} rolling no calibration data, "
                                     f"Please fill in and try again...")
                seg_num = seg_hs
                try:
                    start_index = int(per_img_cali_data[seg_num])
                except BaseException as e:
                    raise ValueError(f"The calibration data [{per_img_cali_data[seg_num]}] formatting error....")
                per_img_roi_data.append([seg_num, start_index])

                img_roi_data.append(per_img_roi_data)
                per_img_roi_data = []
                frame_cnt += 1
    return img_roi_data
